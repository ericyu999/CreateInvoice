#This script is to make ASM worksheet from invoicePackingList excel file
import openpyxl, os
import win32com.client as win32
import time

##########
#Function to delete empty rows at the end of each sheet
def clean_Excel(filename):


    wb = openpyxl.load_workbook(filename)


    for sheet in wb:
        print(sheet.title + ' max row before: ' + str(sheet.max_row))
        for i in range(12, sheet.max_row + 1):

            if sheet['I' + str(i)].value == None:
                sheet.delete_rows(i, sheet.max_row + 1 - i)
        print('max row after: ' + str(sheet.max_row))

    wb.save(filename)

###########
# Function to delete trailing spaces on HS CODE column S
def trim_Excel(filename):

    wb = openpyxl.load_workbook(filename)

    for sheet in wb:
        for row in sheet.iter_rows(min_row=12, min_col=19, max_row=sheet.max_row - 1, max_col=19):
            for cell in row:
                if type(cell.value) == str:
                    try:
                        cell.value = int(cell.value.strip())
                    except ValueError:
                        pass

    wb.save(filename)

#################
# Due to unknown reason, the files created must be open and save by excel before can be uploaded to ASM
# So need to run below function after file created
#def Excel_save_close(filename):
#    excel = win32.gencache.EnsureDispatch('Excel.Application') # opens Excel
#    wb = excel.Workbooks.Open(os.path.join(os.getcwd(), filename))  # opens the file
#    wb.Save()
#    wb.Close()
#    excel.Quit()

##################
# Function to choose the file and save to destFolder
def welcome():
    originFile = input('Please enter the full path of the source excel file (ending with .xlsx): \n')

    destFolder = input('Please enter the folder name you want to save the invoice packing file: \n')

    return originFile, destFolder




def invoice_maker():

    filename, destFolder = welcome()
#    filename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\invoice  packing list 217-22486085 112.xlsx'        old method, replaced by welcome function
    timestr = time.strftime("%Y%m%d-%H%M")
    destFolder = destFolder + '_' + timestr
    savePath = os.path.join(os.environ['USERPROFILE'], 'Documents', 'invoicePackingList', destFolder)

    if os.path.exists(savePath) == False:
        os.makedirs(savePath)
    os.chdir(savePath)

    clean_Excel(filename)
    trim_Excel(filename)

    template_path = os.path.join(os.environ['USERPROFILE'], 'Documents','GitHub','CreateInvoice')
#    templatename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\ASMtemplate.xlsx'        #ASM TEMPLATE FILE LOCATION AND name
    template_name = template_path + '\\invoice & packing list_template.xlsx'


    wb_s = openpyxl.load_workbook(filename, read_only=True, data_only=True)  #source file _s


    newFileList = []

    for sheet in wb_s:
        wb_t = openpyxl.load_workbook(template_name)               #target file _t
        sheet_t = wb_t.active

        newfilename = sheet.title
        newfilename = newfilename.replace(' ', '_')                # replace space in name
        newFileList.append(newfilename + '.xlsx')
        i = 12                                                     # from which row in template to start filling the data
        for eachrow in range (3, sheet.max_row):
            sheet_t.cell(row=i, column=1).value = sheet.cell(row=eachrow, column=4).value   # column A --- D3
            sheet_t.cell(row=i, column=2).value = sheet.cell(row=eachrow, column=6).value   # col B--- F3
            sheet_t.cell(row=i, column=3).value = sheet.cell(row=eachrow, column=8).value  # col C --- H3
            sheet_t.cell(row=i, column=4).value = sheet.cell(row=eachrow, column=12).value   # col D --- L3
            sheet_t.cell(row=i, column=5).value = sheet.cell(row=eachrow, column=14).value   # col E --- N3
            sheet_t.cell(row=i, column=6).value = sheet.cell(row=eachrow, column=15).value   # col F--- O3
            sheet_t.cell(row=i, column=7).value = sheet.cell(row=eachrow, column=16).value   # col G --- P3
            sheet_t.cell(row=i, column=8).value = sheet.cell(row=eachrow, column=17).value    # col H --- Q3
            sheet_t.cell(row=i, column=9).value = sheet.cell(row=eachrow, column=18).value    # col I --- R3
            sheet_t.cell(row=i, column=10).value = sheet.cell(row=eachrow, column=19).value    # col J --- S3
            sheet_t.cell(row=i, column=11).value = sheet.cell(row=eachrow, column=20).value    # col K --- T3
            sheet_t.cell(row=i, column=12).value = sheet.cell(row=eachrow, column=23).value    # col L --- W3


            i += 1



        wb_t.save(newfilename + '.xlsx')

#    for file in newFileList:
#        Excel_save_close(file)


    print("%s files saved successfully %s" %(str(len(wb_s.sheetnames)), savePath))
